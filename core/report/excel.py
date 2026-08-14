"""엑셀 심의자료 생성 — FR-1003-AC1.

이전 구현(12줄)은 워크북을 만들지 않고 `IRR` 결과를 리터럴 딕셔너리로
박아 두고 있었다(`.orch/R13-WP24E-대장.md`). 여기서는 `openpyxl`로
**실제 워크북**을 만들고, 조항이 요구하는 대로 시트를 분리하며(입력·
프로포마·시계열·결과), 주요 계산(현재가치·NPV 합계)은 **셀 수식**으로
쓴다 — 값만 박으면 「기존 엑셀 검토 방식과 병행」(조항 원문)이 성립하지
않는다.

`IRR`은 예외다. spec §15.4 TU-7 결정("수식 포함 XLSX 생성 시 복잡 산식
(IRR·할인회수기간)의 셀 수식 표현 한계 → 표현 불가 항목은 값 + 산식
주석 병기로 대체")에 따라 순환 계산인 IRR은 셀 수식이 아니라 **값 +
셀 주석**으로 담는다.
"""

from __future__ import annotations

import io
from collections.abc import Sequence
from decimal import Decimal
from typing import Any, Final

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.comments import Comment  # type: ignore[import-untyped]

from core.contracts.units import to_won
from core.valuestream.report import STATE_ACCOUNTED, BenefitReport

_SHEET_INPUT: Final[str] = "입력"
_SHEET_PROFORMA: Final[str] = "프로포마"
_SHEET_TIMESERIES: Final[str] = "시계열"
_SHEET_RESULTS: Final[str] = "결과"

_PROFORMA_FIRST_DATA_ROW: Final[int] = 2

#: 편익 계상 내역 하위 표가 시작하는 행. NPV·IRR 두 줄 아래 한 줄을 비운다.
_BENEFIT_FIRST_ROW: Final[int] = 4
_BENEFIT_TABLE_TITLE: Final[str] = "편익 계상 내역 (FR-402-AC6)"
#: `BenefitLine.state` 의 「계상됨」 — 합계 수식의 조건이다. **문면을 여기 베끼는
#: 대신 `core.valuestream.report` 에서 가져온다**(사본을 두면 갈린다).
_BENEFIT_STATE_ACCOUNTED: Final[str] = STATE_ACCOUNTED


def _write_input_sheet(ws: Any, discount_rate: float) -> None:
    ws["A1"] = "할인율"
    ws["B1"] = discount_rate


def _write_proforma_sheet(ws: Any, annual_cashflows_won: Sequence[Decimal | int]) -> int:
    """연도별 현금흐름과 현재가치(셀 수식)를 쓰고, NPV 합계 행 번호를 반환한다."""
    ws["A1"], ws["B1"], ws["C1"] = "연도", "현금흐름(원)", "현재가치(원)"

    for offset, cashflow in enumerate(annual_cashflows_won):
        row = _PROFORMA_FIRST_DATA_ROW + offset
        year = offset + 1
        ws.cell(row=row, column=1, value=year)
        ws.cell(row=row, column=2, value=int(to_won(cashflow)))
        # 현재가치 = 현금흐름 / (1+할인율)^연도 — 입력 시트의 할인율을 참조하는 실제 셀 수식
        ws.cell(row=row, column=3, value=f"=B{row}/(1+{_SHEET_INPUT}!$B$1)^A{row}")

    last_data_row = _PROFORMA_FIRST_DATA_ROW + len(annual_cashflows_won) - 1
    npv_row = last_data_row + 2
    ws.cell(row=npv_row, column=2, value="NPV 합계")
    ws.cell(
        row=npv_row, column=3, value=f"=SUM(C{_PROFORMA_FIRST_DATA_ROW}:C{last_data_row})"
    )
    return npv_row


def _write_timeseries_sheet(ws: Any, hourly_dispatch_kwh: Sequence[float]) -> None:
    ws["A1"], ws["B1"] = "시간(h)", "전력량(kWh)"
    for offset, kwh in enumerate(hourly_dispatch_kwh):
        row = _PROFORMA_FIRST_DATA_ROW + offset
        ws.cell(row=row, column=1, value=offset + 1)
        ws.cell(row=row, column=2, value=kwh)


def _write_results_sheet(
    ws: Any, npv_row: int, irr: float, benefits: BenefitReport | None = None
) -> None:
    ws["A1"], ws["B1"] = "NPV(원)", f"={_SHEET_PROFORMA}!C{npv_row}"

    ws["A2"] = "IRR"
    irr_cell = ws["B2"]
    irr_cell.value = irr
    irr_cell.comment = Comment(
        "spec §15.4 TU-7: IRR은 순환 계산이라 셀 수식으로 표현할 수 없어 "
        "값 + 주석으로 대체한다. 산식: 현금흐름의 NPV가 0이 되는 할인율.",
        "core.report.excel",
    )
    if benefits is not None:
        _write_benefit_breakdown(ws, benefits, first_row=_BENEFIT_FIRST_ROW)


def _write_benefit_breakdown(ws: Any, report: BenefitReport, *, first_row: int) -> None:
    """편익 계상 내역을 **「결과」 시트의 하위 표**로 쓴다 — `FR-402-AC6` / R31 (§7).

    ## 왜 새 시트를 만들지 않는가

    `FR-1003-AC1` 은 시트 구성을 *「입력·프로포마·시계열·결과」* 넷으로 못 박고
    `tests/report/test_report.py::test_excel_sheets_are_separated` 가 그 목록을
    **정확히** 단언한다. **조항 하나를 닫으려고 다른 조항의 계약을 고치지 않는다** —
    R23 이 같은 자리에서 같은 판단을 내렸다(그래서 그 라운드는 XLSX 로 내보내지
    않았다).

    ⚠ **결정 문서 §7 은 「기존 「편익」 시트의 하위 표」라 적었으나 그런 시트는
    없었다.** 시트는 넷뿐이고, 계상 내역은 「무엇을 세었는가」를 설명하므로 **결과**
    시트에 속한다. 실물을 보고 좁힌 것이다.

    ## 네 버킷을 **전부** 쓴다

    「계상됨」만 쓰면 조항이 요구하는 「배타제외·증분만·미화폐화0」이 사라지고, 읽는
    사람은 **세지 않은 편익이 애초에 없었다고** 읽는다 — 그것이 이 조항이 존재하는
    이유다(`payer_gate` 가 「판정 통과」를 명시적으로 돌려주는 것과 같은 근거).
    """
    ws[f"A{first_row}"] = _BENEFIT_TABLE_TITLE
    header_row = first_row + 1
    for column, label in zip("ABCD", ("편익", "지불 주체", "연간액(원)", "상태"), strict=True):
        ws[f"{column}{header_row}"] = label

    row = header_row + 1
    for line in report.all_lines():
        ws[f"A{row}"] = line.name
        ws[f"B{row}"] = line.payer.value
        ws[f"C{row}"] = int(line.annual_value)
        ws[f"D{row}"] = line.state
        row += 1

    # **합계는 셀 수식이다** — 값으로 박으면 「기존 엑셀 검토 방식과 병행」이
    # 성립하지 않는다(이 모듈 머리말과 같은 근거). 계상된 행만 더하도록
    # `SUMIF` 로 상태를 조건에 둔다 — 배타제외·미화폐화0 을 함께 더하면
    # 합계가 `total_accounted()` 와 갈린다.
    if report.all_lines():
        ws[f"A{row}"] = "계상 합계(원)"
        ws[f"C{row}"] = (
            f'=SUMIF(D{header_row + 1}:D{row - 1},"{_BENEFIT_STATE_ACCOUNTED}",'
            f"C{header_row + 1}:C{row - 1})"
        )


def generate_excel(
    annual_cashflows_won: Sequence[Decimal | int],
    discount_rate: float,
    irr: float,
    hourly_dispatch_kwh: Sequence[float] = (),
    benefits: BenefitReport | None = None,
) -> bytes:
    """CBA 결과로 실제 엑셀 워크북 바이트를 만든다 (FR-1003-AC1).

    시트: 입력·프로포마·시계열·결과. NPV는 프로포마 시트의 현재가치
    합계를 참조하는 셀 수식이고, IRR만 TU-7 결정에 따라 값+주석이다.
    호출부가 파일 경로 대신 `tmp_path`나 `io.BytesIO`로 이 바이트를
    받아 검증하면 저장소에 파일이 남지 않는다.
    """
    workbook = Workbook()
    input_ws = workbook.active
    input_ws.title = _SHEET_INPUT
    _write_input_sheet(input_ws, discount_rate)

    proforma_ws = workbook.create_sheet(_SHEET_PROFORMA)
    npv_row = _write_proforma_sheet(proforma_ws, annual_cashflows_won)

    timeseries_ws = workbook.create_sheet(_SHEET_TIMESERIES)
    _write_timeseries_sheet(timeseries_ws, hourly_dispatch_kwh)

    results_ws = workbook.create_sheet(_SHEET_RESULTS)
    _write_results_sheet(results_ws, npv_row, irr, benefits)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
