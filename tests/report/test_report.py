import io
from decimal import Decimal

import openpyxl
import pytest

from core.contracts.assumptions import AssumptionProvider, AssumptionValue
from core.model.model import Model
from core.model.schemas import DERConfig, ModelConfig
from core.report.charts import render_charts
from core.report.excel import generate_excel
from core.report.manifest import create_manifest
from core.report.pdf import generate_pdf
from core.report.sensitivity import rank_influences


@pytest.mark.req("FR-1001-AC1")
def test_influence_ranking_no_input_order() -> None:
    """13.1, 13.2: 영향 인자는 입력 순이 아니라
    민감도 계산에 의한 순위로 제시되어야 함 (FR-1001-AC1).

    영향폭 계산 (metric_fn = 항등 함수):
      B: low=1, high=1000  → delta = |1000 - 1| = 999  (가장 큼)
      A: low=80, high=120  → delta = |120 - 80|  = 40
      C: low=240, high=360 → delta = |360 - 240| = 120
    따라서 B(999) > C(120) > A(40). B 가 1위.
    """

    def identity(x: float) -> float:
        return x

    variables = {
        "A": {"base": 100.0, "low": 80.0, "high": 120.0},
        "B": {"base": 100.0, "low": 1.0, "high": 1000.0},   # B 가 지배적
        "C": {"base": 300.0, "low": 240.0, "high": 360.0},
    }
    ranked = rank_influences(variables, metric_fn=identity)
    assert ranked[0]["name"] == "B"  # B 가 1위로 나와야 함


@pytest.mark.req("FR-1002-AC1", "FR-1002-AC2", "FR-1002-AC4")
def test_sensitivity_thresholds() -> None:
    """13.1: 임계값 산출 및 여유폭 표시 (FR-1002-AC1, AC2, AC4).

    rank_influences 결과에 threshold / margin_pct 필드가 반드시 있어야 한다.
    """
    variables = {
        "price": {"base": 100.0, "low": 50.0, "high": 150.0},
    }
    ranked = rank_influences(variables)
    assert "margin_pct" in ranked[0]
    assert "threshold" in ranked[0]


@pytest.mark.req("FR-1002-AC5", "FR-1002-AC6")
def test_assumption_combinations_and_warnings() -> None:
    """13.3: 가정 결합 표시 및 상단 경고 (FR-1002-AC5, AC6)"""
    # PDF나 리포트에 최악 조건 결합 시 경고가 뜨는지 확인하는 mock
    pdf_content = generate_pdf(has_critical_assumptions=True)
    assert "WARNING" in pdf_content["header"]


@pytest.mark.req("FR-1001-AC2", "FR-1001-AC3", "FR-1001-AC4", "FR-1002-AC3")
def test_formula_triple_representation() -> None:
    """13.4: 산식 3중 표기 및 출처/신뢰도 표시 (FR-1001-AC2~AC4, FR-1002-AC3)"""
    pdf_content = generate_pdf(has_critical_assumptions=False)
    assert "자연어" in pdf_content["formulas"]
    assert "수식" in pdf_content["formulas"]
    assert "대입값" in pdf_content["formulas"]


@pytest.mark.req("FR-1001-AC3")
def test_formula_substituted_value_reflects_actual_inputs() -> None:
    """FR-1001-AC3: 대입값 줄은 «자연어/수식이 갖고 있는 고정 단어」가 아니라
    호출부가 넘긴 실제 편익·비용에서 계산되어야 한다.

    손계산: 편익 1,500,000원 - 비용 900,000원 = NPV 600,000원.
    이전 구현은 인자를 무시하고 항상 "200 = 1000 - 800" 을 반환했으므로,
    이 값을 넘겨 다른 결과가 나오는지 확인하면 동어반복 여부가 드러난다.
    """
    pdf_content = generate_pdf(
        benefit_won=Decimal("1500000"), cost_won=Decimal("900000")
    )
    assert "대입값: 600000 = 1500000 - 900000" in pdf_content["formulas"]


@pytest.mark.req("FR-1001-AC3")
def test_formula_substituted_value_changes_with_different_inputs() -> None:
    """FR-1001-AC3: 편익·비용을 바꾸면 대입값도 함께 바뀌어야 한다.

    손계산: 편익 300원 - 비용 100원 = NPV 200원. 위 테스트의 결과(600,000)와
    달라야 하며, 같다면 인자를 무시하고 고정값을 돌려주는 것이다.
    """
    pdf_content = generate_pdf(benefit_won=Decimal("300"), cost_won=Decimal("100"))
    assert "대입값: 200 = 300 - 100" in pdf_content["formulas"]


@pytest.mark.req("FR-1003-AC2")
def test_pdf_bytes_are_real_reportlab_output() -> None:
    """FR-1003-AC2: `pdf_bytes` 는 reportlab이 실제로 생성한 PDF여야 한다
    (mock 딕셔너리가 아니라 %PDF 로 시작하고 %%EOF 로 끝나는 바이트열).
    저장소에 파일을 남기지 않도록 바이트만 검사한다 (파일 기록 없음).
    """
    pdf_content = generate_pdf(benefit_won=Decimal("1000"), cost_won=Decimal("800"))
    pdf_bytes = pdf_content["pdf_bytes"]
    assert pdf_bytes.startswith(b"%PDF")
    assert pdf_bytes.rstrip().endswith(b"%%EOF")
    assert len(pdf_bytes) > 500  # 16줄 mock 시절엔 PDF 바이트 자체가 없었다


@pytest.mark.req("FR-1002-AC5", "FR-1002-AC6")
def test_pdf_bytes_contain_warning_when_critical() -> None:
    """FR-1002-AC5·AC6: 상단 경고는 반환 dict 뿐 아니라 실제 PDF 산출물에도
    들어가야 한다. 헤더는 영문이라 PDF 콘텐츠 스트림에서 그대로 검색된다.
    """
    critical = generate_pdf(has_critical_assumptions=True)
    normal = generate_pdf(has_critical_assumptions=False)
    assert b"WARNING" in critical["pdf_bytes"]
    assert b"WARNING" not in normal["pdf_bytes"]


@pytest.mark.req("FR-1003-AC1")
def test_excel_sheets_are_separated() -> None:
    """FR-1003-AC1: 「입력·프로포마·시계열·결과 시트를 분리」 요구사항.
    되돌려 받은 바이트를 openpyxl로 다시 읽어 시트 구성을 확인한다
    (구현이 돌려준 dict를 되읽지 않는다 — 그건 이전 mock의 검증 방식이었다).
    """
    xlsx_bytes = generate_excel(
        annual_cashflows_won=[Decimal("1000000"), Decimal("1200000"), Decimal("900000")],
        discount_rate=0.05,
        irr=0.0731,
    )
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    assert wb.sheetnames == ["입력", "프로포마", "시계열", "결과"]


@pytest.mark.req("FR-1003-AC1")
def test_excel_present_value_and_npv_are_real_cell_formulas() -> None:
    """FR-1003-AC1: 「주요 계산은 값이 아닌 셀 수식으로 출력」 요구사항.

    현금흐름 [100만, 120만, 90만원], 3개년이므로 프로포마 데이터 행은
    2~4행(연도 1~3)이고, NPV 합계 행은 그다음 빈 줄을 건너뛴 6행이다
    (손으로 짚은 위치 — _PROFORMA_FIRST_DATA_ROW=2, 행수 3 → last=4, npv=4+2=6).
    각 현재가치 셀은 `=B{row}/(1+입력!$B$1)^A{row}` 꼴의 실제 수식이어야
    하며, 값으로 미리 계산해 두면 안 된다(data_type == 'f').
    """
    xlsx_bytes = generate_excel(
        annual_cashflows_won=[Decimal("1000000"), Decimal("1200000"), Decimal("900000")],
        discount_rate=0.05,
        irr=0.0731,
    )
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    proforma = wb["프로포마"]

    assert proforma["C2"].value == "=B2/(1+입력!$B$1)^A2"
    assert proforma["C3"].value == "=B3/(1+입력!$B$1)^A3"
    assert proforma["C4"].value == "=B4/(1+입력!$B$1)^A4"
    for cell in ("C2", "C3", "C4"):
        assert proforma[cell].data_type == "f", f"{cell} 이 셀 수식이 아닙니다"

    assert proforma["C6"].value == "=SUM(C2:C4)"
    assert proforma["C6"].data_type == "f"

    results = wb["결과"]
    assert results["B1"].value == "=프로포마!C6"
    assert results["B1"].data_type == "f"


@pytest.mark.req("FR-1003-AC1")
def test_excel_irr_is_value_with_comment_per_tu7() -> None:
    """FR-1003-AC1: spec §15.4 TU-7 결정 — IRR은 순환 계산이라 셀 수식 대신
    값 + 주석으로 대체한다. IRR은 셀 수식이면 안 되고(data_type != 'f'),
    넘긴 값이 그대로 담기며, 주석에 TU-7 근거가 남아야 한다.
    """
    xlsx_bytes = generate_excel(
        annual_cashflows_won=[Decimal("1000000")], discount_rate=0.05, irr=0.0731
    )
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    results = wb["결과"]

    assert results["B2"].value == 0.0731
    assert results["B2"].data_type != "f"
    assert results["B2"].comment is not None
    assert "TU-7" in results["B2"].comment.text


class _StubAssumptions(AssumptionProvider):
    """WP-24F 전용 테스트 스텁 — Model 생성에 필요한 vat_rate만 채운다."""

    @property
    def set_name(self) -> str:
        return "wp24f-test"

    @property
    def set_version(self) -> str:
        return "1.0"

    def get(self, key: str) -> AssumptionValue | None:
        if key == "tax.vat_rate":
            return AssumptionValue(
                key=key,
                value=0.1,
                value_unit="소수",
                base_year="2026",
                applicable_scope="",
                derivation_method="가정",
                source="",
                verified_at=None,
                confidence="가정",
            )
        return None


@pytest.mark.req("FR-1003-AC3")
def test_model_to_json_from_json_roundtrip_preserves_scenario() -> None:
    """FR-1003-AC3: 「JSON: 시나리오+전제 정의 전체 (재현용)」.

    이 마커는 원래 엑셀 IRR만 검사하는 테스트에 잘못 붙어 있었다
    (`.orch/R13-WP24E-대장.md` ②). 실제 구현은 `Model.to_json()`/
    `Model.from_json()`(core/model/model.py:54,59)이며, 검증할 것은
    「왕복 후 같은 시나리오가 나오는가」다 — 직렬화 문자열이 비어있지
    않은지만 보면 조항을 검증한 것이 아니다.
    """
    provider = _StubAssumptions()
    config = ModelConfig(
        name="WP-24F 재현성 테스트 모델",
        resources=[
            DERConfig(
                tag="PV",
                params={"name": "PV-A", "capacity_kw": 50.0, "capacity_factor": 0.16},
            )
        ],
    )
    model = Model(config, provider)

    json_str = model.to_json()
    restored = Model.from_json(json_str, provider)

    assert restored.name == "WP-24F 재현성 테스트 모델"
    assert len(restored.resources) == 1
    assert restored.resources[0].name == "PV-A"
    assert restored.resources[0].capacity_kw == 50.0
    # 모델 계층에서 전제(vat_rate)가 재생성 시에도 동일하게 주입되는지 —
    # "전제 정의"가 왕복 후에도 보존됨을 보여준다
    assert restored.resources[0].vat_rate == 0.1


@pytest.mark.req("FR-1005-AC1", "NFR-101-M1")
def test_manifest_reproducibility() -> None:
    """13.6: 실행 매니페스트 재현성 (비트 단위 동일) (FR-1005-AC1, NFR-101-M1)"""
    manifest1 = create_manifest({"seed": 42, "version": "1.0"})
    manifest2 = create_manifest({"seed": 42, "version": "1.0"})
    assert manifest1.hash == manifest2.hash


@pytest.mark.req("FR-1004-AC1")
def test_dashboard_charts() -> None:
    """13.7: 시각화 차트 생성 (FR-1004-AC1)

    **이 테스트는 R16 이전에 아무것도 검증하지 않았다.** `generate_charts()`
    가 돌려주는 dict 에 키 세 개가 있는지만 보았고, 그 값은 영어 문자열
    상수였다 — `"Cumulative Cash Flow Chart with BEP"`. 키 검사는 그 문자열이
    그림인지 설명인지 구분하지 못한다.

    이제 **실제 렌더 바이트**를 본다. 값이 다시 문자열로 돌아가면 여기서
    빨간불이 난다.
    """
    charts = render_charts(
        {
            "cashflows": [-2_000_000.0, 400_000.0, 500_000.0, 600_000.0, 700_000.0],
            "items": {"설비비": 1_500_000.0, "시공비": 400_000.0, "운영비": 100_000.0},
            "influences": [
                {"name": "설비단가", "delta": 800_000.0, "flips_conclusion": True},
                {"name": "할인율", "delta": 250_000.0, "flips_conclusion": False},
            ],
        }
    )
    assert {"cashflow_line", "cost_benefit_pie", "tornado"} <= set(charts)
    for tag, artifact in charts.items():
        assert artifact.mime == "image/png"
        assert artifact.payload.startswith(b"\x89PNG\r\n\x1a\n"), f"{tag}: PNG 가 아니다"


@pytest.mark.manual
@pytest.mark.req("FR-1001-AC5")
def test_manual_check_mc1_stub() -> None:
    """13.8: 수동 검증 MC-1 명세 스텁 (FR-1001-AC5)
    심의 경험자 3명 대상 설명 재현 테스트를 위한 스텁.
    docs/manual-checks.yaml 에 등재되어야 함.
    """
    pass


@pytest.mark.req("FR-1003-AC2")
def test_pdf_report_sections() -> None:
    """13.9: PDF 심의자료 요약 (FR-1003-AC2)"""
    pdf = generate_pdf()
    assert pdf["sections"] == ["cover", "assumptions", "results", "sensitivity", "conclusion"]
