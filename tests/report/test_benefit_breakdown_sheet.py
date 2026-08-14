"""편익 계상 내역이 **리포트에 표시되는가** — `FR-402-AC6` / R31 (결정 §7).

`docs/clause-recheck-2026-08-14.md` 판정: *「4버킷 자료구조는 조항 그대로 있으나
**「리포트에 표시」가 없다.** PDF `_SECTIONS` 에 섹션 없음 · Excel 시트 없음 ·
HTML 에 문자열 0건」*.

## 새 시트를 만들지 않았다

`FR-1003-AC1` 은 시트 구성을 *「입력·프로포마·시계열·결과」* 넷으로 못 박고
`test_excel_sheets_are_separated` 가 그 목록을 **정확히** 단언한다. **조항 하나를
닫으려고 다른 조항의 계약을 고치지 않는다** — R23 이 같은 자리에서 같은 판단을
내렸다.

⚠ **결정 문서 §7 은 「기존 「편익」 시트의 하위 표」라 적었으나 그런 시트는 없었다.**
시트는 넷뿐이다. 계상 내역은 「무엇을 세었는가」를 설명하므로 **결과** 시트에 속한다 —
실물을 보고 좁힌 것이다.

붙드는 것 넷:

    ① 네 버킷이 **전부** 표에 온다        세지 않은 편익이 사라지지 않는다
    ② 상태가 행마다 실린다                「계상됨」과 「배타제외」를 사람이 가른다
    ③ 합계가 **셀 수식**이다              값으로 박으면 병행 검토가 성립하지 않는다
    ④ 시트 목록은 그대로다                다른 조항의 계약을 고치지 않았다

**①이 요점이다.** 「계상됨」만 쓰면 조항이 요구하는 세 버킷이 사라지고, 읽는 사람은
**세지 않은 편익이 애초에 없었다고** 읽는다.
"""

from __future__ import annotations

import io
from decimal import Decimal

import openpyxl  # type: ignore[import-untyped]
import pytest

from core.contracts.units import Money, to_won
from core.contracts.valuestream import Payer
from core.report.excel import generate_excel
from core.valuestream.report import (
    STATE_ACCOUNTED,
    STATE_EXCLUDED,
    STATE_INCREMENT_ONLY,
    STATE_UNMONETIZED_ZERO,
    BenefitLine,
    BenefitReport,
)

_CASHFLOWS = [Decimal("1000000"), Decimal("1200000")]


def _line(name: str, amount: int, state: str) -> BenefitLine:
    return BenefitLine(
        tag=name, name=name, payer=Payer.OPERATOR,
        annual_value=to_won(amount), state=state,
    )


def _report() -> BenefitReport:
    """네 버킷을 모두 채운 리포트 — 버킷 하나만 채우면 ①을 검사할 수 없다."""
    return BenefitReport(
        accounted=[_line("자가소비 절감", 300_000, STATE_ACCOUNTED)],
        excluded=[_line("잉여판매", 200_000, STATE_EXCLUDED)],
        increment_only=[_line("분산편익", 50_000, STATE_INCREMENT_ONLY)],
        unmonetized_zero=[_line("V2G 방전", 0, STATE_UNMONETIZED_ZERO)],
    )


def _results_sheet(benefits: BenefitReport | None):  # type: ignore[no-untyped-def]
    xlsx = generate_excel(
        annual_cashflows_won=_CASHFLOWS, discount_rate=0.05, irr=0.07,
        benefits=benefits,
    )
    return openpyxl.load_workbook(io.BytesIO(xlsx))["결과"]


def _rows(sheet) -> list[tuple[object, ...]]:  # type: ignore[no-untyped-def]
    return [tuple(row) for row in sheet.iter_rows(values_only=True)]


# ── ①② 네 버킷 전부, 상태와 함께 ─────────────────────────────────────

@pytest.mark.req("FR-402-AC6")
def test_every_bucket_reaches_the_sheet_with_its_state() -> None:
    """★★★ 네 버킷이 **전부** 오고 각 행이 자기 상태를 든다.

    「계상됨」만 쓰면 조항이 요구하는 「배타제외·증분만·미화폐화0」이 사라지고,
    읽는 사람은 **세지 않은 편익이 애초에 없었다고** 읽는다 — `payer_gate` 가
    「판정 통과」를 명시적으로 돌려주는 것과 같은 근거다.
    """
    report = _report()
    rows = _rows(_results_sheet(report))

    by_name = {row[0]: row for row in rows if row[0]}
    for line in report.all_lines():
        assert line.name in by_name, f"{line.name}({line.state}) 이 표에 없습니다"
        row = by_name[line.name]
        assert row[1] == line.payer.value, f"{line.name} 의 지불 주체가 없습니다"
        assert row[2] == int(line.annual_value)
        assert row[3] == line.state, (
            f"{line.name} 의 상태가 표에 없습니다 — 「계상됨」과 「배타제외」를 "
            "사람이 가를 수 없습니다"
        )

    # 네 상태가 **모두** 표에 나타난다 — 한 버킷만 그리는 구현을 배제한다
    states = {row[3] for row in rows if row[3]}
    assert {
        STATE_ACCOUNTED, STATE_EXCLUDED, STATE_INCREMENT_ONLY, STATE_UNMONETIZED_ZERO
    } <= states


@pytest.mark.req("FR-402-AC6")
def test_the_table_has_a_title_naming_the_clause() -> None:
    """표에 제목이 있고 그것이 조항을 가리킨다 — 심의자가 근거를 되짚을 수 있게."""
    titles = [row[0] for row in _rows(_results_sheet(_report())) if row[0]]

    assert any("편익 계상 내역" in str(t) for t in titles)
    assert any("FR-402-AC6" in str(t) for t in titles)


# ── ③ 합계가 셀 수식이다 ─────────────────────────────────────────────

@pytest.mark.req("FR-402-AC6", "FR-1003-AC1")
def test_the_accounted_total_is_a_cell_formula_not_a_baked_value() -> None:
    """★★ 합계가 **셀 수식**이다 — 값으로 박으면 병행 검토가 성립하지 않는다.

    이 모듈의 머리말이 *「값만 박으면 「기존 엑셀 검토 방식과 병행」(조항 원문)이
    성립하지 않는다」* 고 적는다. 계상 합계도 같다 — 심의자가 셀을 눌러 무엇을
    더했는지 볼 수 있어야 한다.

    ★ **조건부 합계인 것도 요점이다.** 전부 더하면 배타제외·미화폐화0 이 함께
    들어가 `BenefitReport.total_accounted()` 와 갈리고, 그 어긋남은 두 수가 모두
    그럴듯하므로 드러나지 않는다.
    """
    sheet = _results_sheet(_report())
    formulas = [
        row[2] for row in _rows(sheet)
        if isinstance(row[2], str) and str(row[2]).startswith("=")
    ]

    assert formulas, "계상 합계가 셀 수식이 아닙니다"
    total_formula = next(f for f in formulas if "SUMIF" in str(f))
    assert STATE_ACCOUNTED in str(total_formula), (
        "합계가 상태를 조건에 두지 않습니다 — 배타제외·미화폐화0 이 함께 더해지고 "
        "`total_accounted()` 와 갈립니다"
    )


@pytest.mark.req("FR-402-AC6")
def test_the_formula_condition_matches_the_state_literal_in_the_data() -> None:
    """★★★ 합계 조건 문면이 **행에 실린 상태 문면과 같다.**

    두 문면이 사본이면 한쪽만 고쳐진 상태에서 `SUMIF` 조건이 맞지 않고 **합계가
    조용히 0 이 된다** — 값이 0 이므로 「편익이 없는 사업」으로 그럴듯하게 보인다.
    그래서 R31 이 `STATE_*` 상수를 `core.valuestream.report` 에 세우고 엑셀이
    그것을 가져다 쓰게 했다.
    """
    rows = _rows(_results_sheet(_report()))
    states_in_data = {row[3] for row in rows if row[3]}
    total_formula = next(
        str(row[2]) for row in rows
        if isinstance(row[2], str) and "SUMIF" in str(row[2])
    )

    condition = total_formula.split('"')[1]
    assert condition in states_in_data, (
        f"합계 조건 {condition!r} 이 행의 상태 문면 {sorted(states_in_data)} 에 "
        "없습니다 — 조건이 맞지 않아 합계가 0 이 됩니다"
    )
    assert to_won(sum(int(row[2]) for row in rows
                      if row[3] == condition and isinstance(row[2], int))) == Money(300_000)


# ── ④ 다른 조항의 계약을 고치지 않았다 ───────────────────────────────

@pytest.mark.req("FR-402-AC6", "FR-1003-AC1")
def test_the_sheet_list_is_unchanged() -> None:
    """★ 시트 목록이 그대로다 — 조항 하나를 닫으려고 다른 조항을 열지 않았다.

    `FR-1003-AC1` 이 시트 구성을 넷으로 못 박는다. 계상 내역을 새 시트로 내보내면
    그 계약이 깨지고, 그것은 R23 이 같은 자리에서 하지 않기로 한 일이다.
    """
    xlsx = generate_excel(
        annual_cashflows_won=_CASHFLOWS, discount_rate=0.05, irr=0.07,
        benefits=_report(),
    )
    workbook = openpyxl.load_workbook(io.BytesIO(xlsx))

    assert workbook.sheetnames == ["입력", "프로포마", "시계열", "결과"]


@pytest.mark.req("FR-1003-AC1")
def test_omitting_the_breakdown_leaves_the_results_sheet_as_before() -> None:
    """내역을 주지 않으면 종전과 같다 — 기존 호출부를 깨지 않았다.

    `benefits` 를 필수로 만들면 이 조항을 모르는 호출부가 전부 깨지고, 그때
    사람은 급히 빈 리포트를 넘기게 된다 — 빈 표는 「편익이 없다」로 읽힌다.
    """
    rows = _rows(_results_sheet(None))
    labels = [row[0] for row in rows if row[0]]

    assert labels == ["NPV(원)", "IRR"]
